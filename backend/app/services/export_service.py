import csv
import io

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer

from app.ai.preprocessors.text_cleaner import strip_html_for_display
from app.ai.rules.keywords import get_top_words
from app.schemas.analysis import SentimentSummary

_SENTIMENT_COLORS = {"Positive": "#22C55E", "Negative": "#EF4444", "Neutral": "#9CA3AF"}


def build_csv_export(job, rows: list[tuple]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["author", "comment", "like_count", "published_at", "sentiment", "confidence"])
    for comment, result in rows:
        writer.writerow(
            [
                comment.author_display_name,
                strip_html_for_display(comment.text),
                comment.like_count,
                comment.published_at,
                result.label if result else "",
                f"{result.confidence:.4f}" if result and result.confidence is not None else "",
            ]
        )
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def _sentiment_pie_png(summary: SentimentSummary) -> io.BytesIO:
    labels = ["Positive", "Negative", "Neutral"]
    values = [summary.positive_count, summary.negative_count, summary.neutral_count]
    fig, ax = plt.subplots(figsize=(5, 5))
    ax.pie(
        values,
        labels=labels,
        autopct="%1.1f%%",
        colors=[_SENTIMENT_COLORS[label] for label in labels],
        textprops={"color": "#1f2937"},
    )
    ax.set_title("Sentiment Distribution")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _sentiment_bar_png(summary: SentimentSummary) -> io.BytesIO:
    labels = ["Positive", "Negative", "Neutral"]
    values = [summary.positive_count, summary.negative_count, summary.neutral_count]
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.bar(labels, values, color=[_SENTIMENT_COLORS[label] for label in labels])
    ax.set_ylabel("Comments")
    ax.set_title("Comments by Sentiment")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def _language_pie_png(summary: SentimentSummary) -> io.BytesIO:
    labels = ["Arabic", "English", "Mixed", "Other"]
    values = [summary.arabic_count, summary.english_count, summary.mixed_count, summary.other_count]
    labels = [label for label, v in zip(labels, values, strict=True) if v > 0]
    values = [v for v in values if v > 0]
    fig, ax = plt.subplots(figsize=(5, 5))
    if values:
        ax.pie(values, labels=labels, autopct="%1.1f%%")
    ax.set_title("Language Distribution")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


def build_pdf_export(job, summary: SentimentSummary, insights: list[str]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(
        Paragraph(f"YouTube Analysis Report: {job.video_title or job.video_url}", styles["Title"])
    )
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Sentiment Distribution", styles["Heading2"]))
    elements.append(Image(_sentiment_pie_png(summary), width=300, height=300))

    total = summary.positive_count + summary.negative_count + summary.neutral_count
    for label, count in [
        ("Positive", summary.positive_count),
        ("Negative", summary.negative_count),
        ("Neutral", summary.neutral_count),
    ]:
        pct = round(count / total * 100, 1) if total else 0
        elements.append(Paragraph(f"{label} Comments: {pct}% ({count})", styles["BodyText"]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("AI Insights", styles["Heading2"]))
    for insight in insights:
        elements.append(Paragraph(f"• {insight}", styles["BodyText"]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Sentiment Comparison", styles["Heading2"]))
    elements.append(Image(_sentiment_bar_png(summary), width=400, height=270))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Language Distribution", styles["Heading2"]))
    elements.append(Image(_language_pie_png(summary), width=300, height=300))

    doc.build(elements)
    return buffer.getvalue()


def build_excel_export(
    job,
    summary: SentimentSummary,
    rows: list[tuple],
) -> bytes:
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    total = summary.positive_count + summary.negative_count + summary.neutral_count
    summary_rows = [
        ("Video", job.video_title or job.video_url),
        ("Total Comments", total),
        ("Positive", summary.positive_count),
        ("Positive %", round(summary.positive_count / total * 100, 1) if total else 0),
        ("Negative", summary.negative_count),
        ("Negative %", round(summary.negative_count / total * 100, 1) if total else 0),
        ("Neutral", summary.neutral_count),
        ("Neutral %", round(summary.neutral_count / total * 100, 1) if total else 0),
        ("Spam", summary.spam_count),
        ("Sarcasm", summary.sarcasm_count),
        ("Avg Confidence", f"{(summary.average_confidence or 0):.0%}"),
        ("Arabic", summary.arabic_count),
        ("English", summary.english_count),
        ("Mixed", summary.mixed_count),
        ("Other", summary.other_count),
    ]
    for r, (key, value) in enumerate(summary_rows, start=1):
        summary_ws.cell(row=r, column=1, value=key)
        summary_ws.cell(row=r, column=2, value=value)
    summary_ws.column_dimensions["A"].width = 20
    summary_ws.column_dimensions["B"].width = 40

    header = ["Author", "Comment", "Likes", "Published At", "Sentiment", "Confidence", "Spam", "Sarcasm"]

    def write_sheet(title: str, sheet_rows: list[tuple]):
        ws = wb.create_sheet(title)
        ws.append(header)
        for comment, sentiment_result, spam_result, sarcasm_result in sheet_rows:
            ws.append(
                [
                    comment.author_display_name,
                    strip_html_for_display(comment.text),
                    comment.like_count,
                    comment.published_at,
                    sentiment_result.label if sentiment_result else "",
                    sentiment_result.confidence
                    if sentiment_result and sentiment_result.confidence is not None
                    else "",
                    "Yes" if spam_result and spam_result.label == "spam" else "No",
                    "Yes" if sarcasm_result and sarcasm_result.label == "sarcastic" else "No",
                ]
            )
        for col_idx, col_name in enumerate(header, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_name != "Comment" else 60

    write_sheet("All Comments", rows)
    write_sheet("Positive", [r for r in rows if r[1] and r[1].label == "positive"])
    write_sheet("Negative", [r for r in rows if r[1] and r[1].label == "negative"])
    write_sheet("Neutral", [r for r in rows if r[1] and r[1].label == "neutral"])
    write_sheet("Spam", [r for r in rows if r[2] and r[2].label == "spam"])

    kw_ws = wb.create_sheet("Keywords")
    pos_words = get_top_words([r[0].text for r in rows if r[1] and r[1].label == "positive"], 20)
    neg_words = get_top_words([r[0].text for r in rows if r[1] and r[1].label == "negative"], 20)
    kw_ws.append(["Positive Keyword", "Count", "Negative Keyword", "Count"])
    for i in range(max(len(pos_words), len(neg_words))):
        pw, pc = pos_words[i] if i < len(pos_words) else ("", "")
        nw, nc = neg_words[i] if i < len(neg_words) else ("", "")
        kw_ws.append([pw, pc, nw, nc])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
